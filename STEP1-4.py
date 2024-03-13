#######Automatic Program########


###STEP1
import pandas as pd
import numpy as np  # Ensure numpy is imported
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Read the Excel file and sort it by 'Start'
Input = pd.read_excel("E:/My_python_work/LG4processed.xlsx", sheet_name='LG4')
Input = Input.sort_values(by='POS-2').reset_index(drop=True)

# Define a function to calculate the counts of A, H, B, - at once
def count_symbols(row):
    counts = (row[7:] == 'A').sum(), (row[7:] == 'H').sum(), (row[7:] == 'B').sum(), (row[7:] == '-').sum()
    return counts

# Apply the function and add the results to new columns
Input[['A_Count', 'H_Count', 'B_Count', '-_Count']] = Input.apply(lambda row: pd.Series(count_symbols(row)), axis=1)

# Define functions to calculate f(A) and f(H)
def calculate_f_a(row):
    A, H, B = row['A_Count'], row['H_Count'], row['B_Count']
    denominator = 2 * (A + B + H)
    return (2 * A + H) / denominator if denominator != 0 else np.nan

def calculate_f_h(row):
    A, H, B = row['A_Count'], row['H_Count'], row['B_Count']
    denominator = A + B + H
    return H / denominator if denominator != 0 else np.nan

# Calculate f(A) and f(H)
Input['f(A)'] = Input.apply(calculate_f_a, axis=1)
Input['f(H)'] = Input.apply(calculate_f_h, axis=1)

# Save the processed data to a new Excel sheet
wb = load_workbook("E:/My_python_work/LG4processed.xlsx")
ws = wb.create_sheet("LG4-step1")
for r in dataframe_to_rows(Input, index=False, header=True):
    ws.append(r)
wb.save('E:/My_python_work/LG4processed.xlsx')

###STEP2
import pandas as pd

# Load the specified worksheet
file_path = "E:/My_python_work/LG4processed.xlsx"
sheet_name = "LG4-step1"
df = pd.read_excel(file_path, sheet_name=sheet_name)

# Find the index of the column "A_Count"
a_count_index = df.columns.get_loc("A_Count")

# Initialize a list of indexes for rows to be deleted
rows_to_drop = []

# Iterate through each row starting from the first row
for index, row in df.iloc[0:].iterrows():
    # Calculate the count of "-" and the total number of cells (from the 6th column to the column before "A_Count")
    total_cells = a_count_index - 6  # Starting from the 6th column
    dash_count = (row[5:a_count_index] == "-").sum()
    
    # Check if the count of "-" is 20% or more of the total
    if dash_count / total_cells >= 0.2:
        # Mark this row for deletion
        rows_to_drop.append(index)

# Delete the marked rows
df_cleaned = df.drop(rows_to_drop)

# Initialize a new list to store indexes of rows to be deleted
rows_to_drop = []

# Iterate through each row from the first row, checking if it meets the deletion criteria
for index, row in df_cleaned.iloc[0:].iterrows():
    f_a = row['f(A)']
    f_h = row['f(H)']
    
    # Check if any of the deletion criteria are met
    if (f_a == 0 and f_h == 0) or (f_a == 1 and f_h == 0) or f_a > 0.9 or f_a < 0.1 or f_h < 0.15:
        rows_to_drop.append(index)

# Delete the marked rows
df_final = df_cleaned.drop(rows_to_drop)
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill

# Load the existing Excel file
file_path = "E:/My_python_work/LG4processed.xlsx"
wb = load_workbook(file_path)

# Create a new worksheet "Segregating loci"
ws = wb.create_sheet("Segregating loci")

# Write the data from df_final into the new worksheet row by row, cell by cell
for r_idx, row in enumerate(dataframe_to_rows(df_final, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)
        # Set cell format as needed here
        # For example, set cell color based on value
        if value == 'A':
            cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        elif value == 'B':
            cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')  # Light yellow
        elif value == 'H':
            cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  
        # More conditions and formatting can be added as needed

# Save the workbook
wb.save(file_path)

###STEP3

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill

# Step 1: Read and filter data
file_path = "E:/My_python_work_xibanya/LG4processed.xlsx"
sheet_name = "Segregating loci"
df = pd.read_excel(file_path, sheet_name=sheet_name)

# Keep only the rows where A_Count=0 or B_Count=0
df_filtered = df[(df['A_Count'] == 0) | (df['B_Count'] == 0)]

# Step 2: Save the filtered data
wb = load_workbook(file_path)
ws = wb.create_sheet("1 to 1")  # Create a new worksheet "1 to 1"

# Write the filtered data into the new worksheet row by row, cell by cell
for r_idx, row in enumerate(dataframe_to_rows(df_filtered, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)
        # Step 3: Apply cell color
        if value == 'A':
            cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # Light red
        elif value == 'B':
            cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')  # Light yellow
        elif value == 'H':
            cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Light green

# Step 4: Save the workbook
wb.save(file_path)

# Repeat for different criteria and worksheets as per your code blocks, such as for "A-B" worksheet and "1 to 2 to 1" worksheet, and also for creating "Parent1" and "Parent2" worksheets. Adjust the specific steps (like the filtering conditions and sheet names) according to each block.

# Here's a template for those repetitions:
# Step 1: Read and filter data
# Keep only rows where H_Count=0 and f(A) is greater than 0.44 and less than 0.56 (or other conditions based on the block)
# Step 2: Save the filtered data
# Step 3: Apply cell color (similarly applying specific colors based on cell values)
# Step 4: Save the workbook

# For creating Parent1 and Parent2 worksheets:
# Load data
# Filter rows based on specific genetic conditions
# Perform replace operations (e.g., replacing all 'B' with 'A' from the 6th column onwards)
# Save the filtered data in new worksheets named "Parent1" and "Parent2"
# Apply cell color as needed
# Save the workbook

###STEP4-PART1

import pandas as pd
from openpyxl import load_workbook

# Load data
file_path = "E:/My_python_work_xibanya/LG4processed.xlsx"
sheet_name = "Parent1"
df = pd.read_excel(file_path, sheet_name=sheet_name)

# Sort by the first column in ascending order
df_sorted = df.sort_values(by=df.columns[0], ascending=True)

import pandas as pd

# Specify the file path
file_path = "E:/My_python_work_xibanya/LG4processed.xlsx"

# Load the 'Parent1' worksheet
df = pd.read_excel(file_path, sheet_name='Parent1')

# Proceed with your data processing...
# For example, sort by the first column in ascending order
df_sorted = df.sort_values(by=df.columns[0])

# Calculate the frequency of each chromosome marker in the second column
chromosome_counts = df_sorted.iloc[:, 1].value_counts()
# Find the most common chromosome marker
most_common_chromosome = chromosome_counts.idxmax()

# Filter out rows not belonging to the most frequently occurring chromosome marker
df_not_most_common = df_sorted[df_sorted.iloc[:, 1] != most_common_chromosome]

# Keep the remaining rows (those belonging to the most frequently occurring chromosome marker) in the original DataFrame
df_sorted = df_sorted[df_sorted.iloc[:, 1] == most_common_chromosome]

# Analyze the frequency of 'A' and 'H' in the first 100 rows of the sixth column
letter_counts = df_sorted.iloc[:100, 5].value_counts()  # Assuming the index of the sixth column is 5

# Default sorting for when A is the majority
sort_ascending = True  # True when A is majority, False when H is majority

# Determine which letter is the majority and set the corresponding sorting strategy
if letter_counts.get('H', 0) > letter_counts.get('A', 0):
    sort_ascending = False

# Perform the sorting
# If A is the majority, then sort the sixth column in ascending order, and the first column in ascending order
# If H is the majority, then sort the sixth column in descending order, and the first column in ascending order
df_sorted = df_sorted.sort_values(by=[df_sorted.columns[5], df_sorted.columns[0]], 
                                  ascending=[sort_ascending, True])
# Step 1: Find the start row where the value no longer increases
df=df_sorted
start_row = 0
for i in range(1, len(df)):
    if df.iloc[i, 0] < df.iloc[i-1, 0]:
        start_row = i
        break

# If no disruption in the increasing order is found, start_row will remain 0, meaning no replacement operation is needed
if start_row > 0:
    # Step 2: Perform letter replacement
    for i in range(start_row, len(df)):
        for j in range(5, df.shape[1]):  # Starting from the sixth column
            cell_value = df.iloc[i, j]
            if cell_value == 'A':
                df.iloc[i, j] = 'K'  # First, replace all 'A's with 'K's
            elif cell_value == 'H':
                df.iloc[i, j] = 'A'  # Then, replace all 'H's with 'A's

    # Now replace all 'K's with 'H's
    for i in range(start_row, len(df)):
        for j in range(5, df.shape[1]):  # Starting from the sixth column
            if df.iloc[i, j] == 'K':
                df.iloc[i, j] = 'H'

# Re-sort
df = df.sort_values(by=df.columns[0])

# Delete old count columns and function columns (if they exist)
for col in ['A_Count', 'B_Count', 'H_Count', '-_Count', 'f(A)', 'f(H)']:
    if col in df.columns:
        df.drop(col, axis=1, inplace=True)

# Recalculate count columns
# Define a function to count
def count_letters_from_sixth_column(row, letter):
    return (row.iloc[5:] == letter).sum()  # Count starting from the sixth column

df['A_Count'] = df.apply(lambda row: count_letters_from_sixth_column(row, 'A'), axis=1)
df['B_Count'] = df.apply(lambda row: count_letters_from_sixth_column(row, 'B'), axis=1)
df['H_Count'] = df.apply(lambda row: count_letters_from_sixth_column(row, 'H'), axis=1)
df['-_Count'] = df.apply(lambda row: count_letters_from_sixth_column(row, '-'), axis=1)

# Recalculate f(A) and f(H)
df['f(A)'] = df.apply(lambda row: (2 * row['A_Count'] + row['H_Count']) / (2 * (row['A_Count'] + row['B_Count'] + row['H_Count'])) if (row['A_Count'] + row['B_Count'] + row['H_Count']) > 0 else np.nan, axis=1)
df['f(H)'] = df.apply(lambda row: row['H_Count'] / (row['A_Count'] + row['B_Count'] + row['H_Count']) if (row['A_Count'] + row['B_Count'] + row['H_Count']) > 0 else np.nan, axis=1)

# Save to new sheets "Parent1-1" and "DSC.1"
file_path = "E:/My_python_work_xibanya/LG4processed.xlsx"
with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    df.to_excel(writer, sheet_name="Parent1-1", index=False)
    df_not_most_common.to_excel(writer, sheet_name="DSC.1", index=False)

file_path = "E:/My_python_work_xibanya/LG4processed.xlsx"
sheet_name = "Parent1-1"
df = pd.read_excel(file_path, sheet_name=sheet_name)

# Assuming df is the DataFrame already loaded and each cell contains only the letters A or H

def compute_difference(row1, row2):
    """Calculate the difference between two rows"""
    return sum(a != b for a, b in zip(row1, row2))

def flip_letters(s):
    """Flip A and H in the row"""
    return s.replace({'A': 'H', 'H': 'A'})

# Initialize a list to save the optimized rows
optimized_rows = [df.iloc[0]]

for i in range(1, len(df)):
    current_row = df.iloc[i]
    prev_row = optimized_rows[-1]
    
    # Calculate the difference between the current row and the previous row
    diff_with_current = compute_difference(prev_row[5:], current_row[5:])
    
    # Calculate the difference if the current row is flipped, compared to the previous row
    flipped_current_row = flip_letters(current_row[5:])
    diff_with_flipped = compute_difference(prev_row[5:], flipped_current_row)
    
    # If flipping reduces the difference, save the flipped row
    if diff_with_flipped < diff_with_current:
        optimized_rows.append(current_row.apply(lambda x: 'H' if x == 'A' else 'A' if x == 'H' else x))
    else:
        optimized_rows.append(current_row)

# Combine the optimized rows into a new DataFrame
optimized_df = pd.DataFrame(optimized_rows).reset_index(drop=True)

# Save the optimized DataFrame to a new sheet "Parent1-2"
with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    optimized_df.to_excel(writer, sheet_name="Parent1-2", index=False)

# Assuming df is your DataFrame, and each cell only contains 'A' or 'H'

def similarity(row1, row2):
    """Calculate the similarity between two rows"""
    return sum(a == b for a, b in zip(row1, row2))

def flip_row(row):
    """Swap A and H in the row"""
    return ['H' if cell == 'A' else 'A' if cell == 'H' else cell for cell in row]

# Initialize the result DataFrame, keeping the first row unchanged
optimized_rows = [df.iloc[0, 5:].tolist()]

# Iterate through the rows of the DataFrame, comparing the similarity between adjacent rows
for i in range(1, len(df)):
    current_row = df.iloc[i, 5:].tolist()
    prev_optimized_row = optimized_rows[-1]
    
    # Calculate the similarity without flipping
    sim_without_flip = similarity(prev_optimized_row, current_row)
    # Calculate the similarity after flipping
    flipped_row = flip_row(current_row)
    sim_with_flip = similarity(prev_optimized_row, flipped_row)
    
    # Decide whether to flip based on whether the similarity improves
    if sim_with_flip > sim_without_flip:
        optimized_rows.append(flipped_row)
    else:
        optimized_rows.append(current_row)

# Update the DataFrame with the optimized row data (only updating the sixth column and onwards)
for i, row in enumerate(optimized_rows, start=0):
    df.iloc[i, 5:] = row

# Save the DataFrame with optimized rows to a new sheet 'Parent1-3'
with pd.ExcelWriter("E:/My_python_work_xibanya/LG4processed.xlsx", engine='openpyxl', mode='a') as writer:
    df.to_excel(writer, sheet_name='Parent1-3', index=False)


###STEP4-PART2
import pandas as pd
from openpyxl import load_workbook

# Load data
file_path = "E:/My_python_work_xibanya/LG4processed-TEST-error.xlsx"  # Corrected the path
sheet_name = "Parent1-3"
df = pd.read_excel(file_path, sheet_name=sheet_name)

# Calculate the number of cells from the sixth column in the first row that meet the conditions 'A', 'B', 'H', '-'
total_cells = df.iloc[0, 5:].apply(lambda x: x in ['A', 'B', 'H', '-']).sum()

final_df = df

def find_bounds_for_all_rows(df):
    bounds = {}
    for index, row in df.iterrows():
        pos = row['POS-2']
        # Upper bound: Current position minus 102400
        upper_bound = df[df['POS-2'] <= pos - 102400].index.max() if not df[df['POS-2'] <= pos - 102400].empty else None
        # Lower bound: Current position plus 102400
        lower_bound = df[df['POS-2'] >= pos + 102400].index.min() if not df[df['POS-2'] >= pos + 102400].empty else None
        bounds[index] = (upper_bound, lower_bound)
    return bounds

def validate_rows_with_bounds(df, bounds):
    error_counts = []
    for index, row in df.iterrows():
        error_count = 0
        upper_bound_index, lower_bound_index = bounds.get(index, (None, None))
        if upper_bound_index is None or lower_bound_index is None:
            error_counts.append(error_count)
            continue
        for col in range(5, df.shape[1] - 4):  # Ignoring the last 4 special columns
            cell = row.iloc[col]
            if cell == '-':
                continue
            upper_cell = df.iloc[upper_bound_index, col] if upper_bound_index else None
            lower_cell = df.iloc[lower_bound_index, col] if lower_bound_index else None
            if upper_cell == '-' or lower_cell == '-':
                continue
            if cell != upper_cell and cell != lower_cell:
                error_count += 1
        error_counts.append(error_count)
    return error_counts

# Pre-calculate the upper and lower bound indices for each row
bounds = find_bounds_for_all_rows(final_df)

# Use the pre-calculated indices to count errors
error_counts = validate_rows_with_bounds(final_df, bounds)

# Add the error counts to the DataFrame
final_df['Error_Count'] = error_counts
def find_bounds_for_all_rows_avoiding_errors(df, bounds, error_threshold=10):
    new_bounds = {}
    for index in bounds:
        # Get the original upper and lower bound indices
        upper_bound_index, lower_bound_index = bounds[index]

        # Avoid rows with Error_Count > 10 and find a new upper bound
        while upper_bound_index is not None and df.at[upper_bound_index, 'Error_Count'] > error_threshold:
            upper_bound_index -= 1  # Move upwards
            if upper_bound_index < 0:
                upper_bound_index = None
                break

        # Avoid rows with Error_Count > 10 and find a new lower bound
        while lower_bound_index is not None and df.at[lower_bound_index, 'Error_Count'] > error_threshold:
            lower_bound_index += 1  # Move downwards
            if lower_bound_index >= len(df):
                lower_bound_index = None
                break

        new_bounds[index] = (upper_bound_index, lower_bound_index)
    return new_bounds

def reevaluate_error_counts(df, bounds):
    error_counts = []
    for index, row in df.iterrows():
        upper_bound_index, lower_bound_index = bounds.get(index, (None, None))
        error_count = 0
        if upper_bound_index is None or lower_bound_index is None:
            error_counts.append(error_count)
            continue
        for col in range(5, df.shape[1] - 4):
            cell = row.iloc[col]
            if cell == '-' or df.iloc[upper_bound_index, col] == '-' or df.iloc[lower_bound_index, col] == '-':
                continue
            if cell != df.iloc[upper_bound_index, col] and cell != df.iloc[lower_bound_index, col]:
                error_count += 1
        error_counts.append(error_count)
    return error_counts

# Identify all row indices with Error_Count greater than 10
high_error_indices = final_df[final_df['Error_Count'] > 10].index

# Pre-calculate upper and lower bound indices for each row, avoiding rows with Error_Count > 10
bounds = find_bounds_for_all_rows(final_df)
new_bounds = find_bounds_for_all_rows_avoiding_errors(final_df, bounds)

# Reevaluate error counts using the new upper and lower bound indices
new_error_counts = reevaluate_error_counts(final_df, new_bounds)

# Update the Error_Count column in the DataFrame
for index, error_count in zip(final_df.index, new_error_counts):
    final_df.at[index, 'Error_Count'] = error_count

# After the processing logic is complete, you can continue to work with df as needed, such as removing rows where Error_Count still exceeds 10
#save
with pd.ExcelWriter("E:/My_python_work_xibanya//LG4processed-TEST-erro.xlsx", engine='openpyxl', mode='a') as writer:
    final_df.to_excel(writer, sheet_name='Parent1-5', index=False)

import pandas as pd
from openpyxl import load_workbook

# Load data
file_path = "E:/My_python_work_xibanya//LG4processed-TEST-erro.xlsx"
sheet_name = "Parent1-3"
df = pd.read_excel(file_path, sheet_name=sheet_name)

# Calculate the number of cells from the sixth column onwards that meet the conditions 'A', 'B', 'H', '-' in the first row
total_cells = df.iloc[0, 5:].apply(lambda x: x in ['A', 'B', 'H', '-']).sum()

final_df = df

def find_bounds_for_all_rows(df):
    bounds = {}
    for index, row in df.iterrows():
        pos = row['POS-2']
        # Upper bound: the current position minus 102400
        upper_bound = df[df['POS-2'] <= pos - 102400].index.max() if not df[df['POS-2'] <= pos - 102400].empty else None
        # Lower bound: the current position plus 102400
        lower_bound = df[df['POS-2'] >= pos + 102400].index.min() if not df[df['POS-2'] >= pos + 102400].empty else None
        bounds[index] = (upper_bound, lower_bound)
    return bounds

def validate_rows_with_bounds(df, bounds):
    error_counts = []
    for index, row in df.iterrows():
        error_count = 0
        upper_bound_index, lower_bound_index = bounds.get(index, (None, None))
        if upper_bound_index is None or lower_bound_index is None:
            error_counts.append(error_count)
            continue
        for col in range(5, df.shape[1] - 4):  # Ignoring the last 4 special columns
            cell = row.iloc[col]
            if cell == '-':
                continue
            upper_cell = df.iloc[upper_bound_index, col] if upper_bound_index else None
            lower_cell = df.iloc[lower_bound_index, col] if lower_bound_index else None
            if upper_cell == '-' or lower_cell == '-':
                continue
            if cell != upper_cell and cell != lower_cell:
                error_count += 1
        error_counts.append(error_count)
    return error_counts

# Pre-calculate the upper and lower bounds index for each row
bounds = find_bounds_for_all_rows(final_df)

# Count errors using pre-calculated indexes
error_counts = validate_rows_with_bounds(final_df, bounds)

# Add the error counts to DataFrame
final_df['Error_Count2'] = error_counts

def find_bounds_for_all_rows_avoiding_errors(df, bounds, error_threshold=10):
    new_bounds = {}
    for index in bounds:
        # Get the original upper and lower bound indexes
        upper_bound_index, lower_bound_index = bounds[index]

        # Avoid rows with Error_Count > 10 to find a new upper bound
        while upper_bound_index is not None and df.at[upper_bound_index, 'Error_Count2'] > error_threshold:
            upper_bound_index -= 1  # Move up
            if upper_bound_index < 0:
                upper_bound_index = None
                break

        # Avoid rows with Error_Count > 10 to find a new lower bound
        while lower_bound_index is not None and df.at[lower_bound_index, 'Error_Count2'] > error_threshold:
            lower_bound_index += 1  # Move down
            if lower_bound_index >= len(df):
                lower_bound_index = None
                break

        new_bounds[index] = (upper_bound_index, lower_bound_index)
    return new_bounds

def reevaluate_error_counts(df, bounds):
    error_counts = []
    for index, row in df.iterrows():
        upper_bound_index, lower_bound_index = bounds.get(index, (None, None))
        error_count = 0
        if upper_bound_index is None or lower_bound_index is None:
            error_counts.append(error_count)
            continue
        for col in range(5, df.shape[1] - 4):
            cell = row.iloc[col]
            if cell == '-' or df.iloc[upper_bound_index, col] == '-' or df.iloc[lower_bound_index, col] == '-':
                continue
            if cell != df.iloc[upper_bound_index, col] and cell != df.iloc[lower_bound_index, col]:
                error_count += 1
        error_counts.append(error_count)
    return error_counts

# Identify all row indexes with Error_Count greater than 10
high_error_indices = final_df[final_df['Error_Count2'] > 10].index

# Pre-calculate the upper and lower bounds index for each row, avoiding rows with Error_Count > 10
bounds = find_bounds_for_all_rows(final_df)
new_bounds = find_bounds_for_all_rows_avoiding_errors(final_df, bounds)

# Count errors using the new upper and lower bounds indexes
new_error_counts = reevaluate_error_counts(final_df, new_bounds)

# Update the Error_Count column in DataFrame
for index, error_count in zip(final_df.index, new_error_counts):
    final_df.at[index, 'Error_Count2'] = error_count

# After completing the logic processing, you can continue to process df as needed, such as deleting rows where Error_Count still exceeds 10
#Save
with pd.ExcelWriter("E:/My_python_work_xibanya//LG4processed-TEST-erro.xlsx", engine='openpyxl', mode='a') as writer:
    final_df.to_excel(writer, sheet_name='Parent1-6', index=False)

import pandas as pd

# Step 1: Load datasets
df_parent1_5 = pd.read_excel('E:/My_python_work_xibanya//LG4processed-TEST-erro.xlsx', sheet_name='Parent1-5')
df_parent1_6 = pd.read_excel('E:/My_python_work_xibanya//LG4processed-TEST-erro.xlsx', sheet_name='Parent1-6')

# Assuming df_parent1_5 and df_parent1_6 are already sorted in the same order, or they have a common key for merging
# If not, you might need to merge them based on a key
df_combined = pd.concat([df_parent1_5, df_parent1_6[['Error_Count2']]], axis=1)

# Step 3: Process data
rows_to_delete = []
final_rows = []

for index, row in df_combined.iterrows():
    if row['Error_Count'] > 10 and row['Error_Count2'] > 10:
        # If both error counts are greater than 10, record the row index for deletion
        rows_to_delete.append(index)
    elif row['Error_Count'] <= 10 or row['Error_Count2'] <= 10:
        # Keep the row with the lesser error count
        row_to_keep = row.drop('Error_Count2') if row['Error_Count'] <= row['Error_Count2'] else df_parent1_6.loc[index]
        final_rows.append(row_to_keep)

# Create the final DataFrame and a DataFrame of deleted rows
df_final = pd.DataFrame(final_rows)
df_deleted = df_combined.loc[rows_to_delete]

# Step 4: Save results
with pd.ExcelWriter('E:/My_python_work_xibanya//LG4processed-TEST-erro.xlsx', engine='openpyxl', mode='a') as writer:
    df_final.to_excel(writer, sheet_name='Parent1-7', index=False)
    df_deleted.to_excel(writer, sheet_name='DSC1.1', index=False)
